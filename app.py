import streamlit as st
import pandas as pd
import io
import re
import math
import json
import uuid
from collections import Counter
import validator
import os
import streamlit.components.v1 as components

# ------------------------------------------------------------
# STREAMLIT SETUP
# ------------------------------------------------------------

def _get_build_info():
    """
    Best-effort build identifier for Streamlit Cloud debugging.
    Tries common env vars first, then falls back to local git metadata if available.
    """
    for k in ("GIT_SHA", "GITHUB_SHA", "COMMIT_SHA", "SOURCE_VERSION"):
        v = os.environ.get(k)
        if v:
            return v[:12]

    try:
        head_path = os.path.join(os.path.dirname(__file__), ".git", "HEAD")
        if not os.path.exists(head_path):
            return "unknown"
        with open(head_path, "r", encoding="utf-8") as f:
            head = f.read().strip()
        if head.startswith("ref:"):
            ref = head.split(" ", 1)[1].strip()
            ref_path = os.path.join(os.path.dirname(__file__), ".git", *ref.split("/"))
            if os.path.exists(ref_path):
                with open(ref_path, "r", encoding="utf-8") as rf:
                    sha = rf.read().strip()
                    return sha[:12] if sha else "unknown"
        return head[:12] if head else "unknown"
    except Exception:
        return "unknown"


st.set_page_config(page_title="Diamond Inventory Validator", layout="wide")
st.title("Diamond & Lab-Grown Inventory Validator")
st.markdown("""
Upload your **supplier inventory file** to validate against the internal rule set.
""")
st.caption(f"Build: `{_get_build_info()}`")

# ------------------------------------------------------------
# UI HELPERS
# ------------------------------------------------------------

def render_copy_to_clipboard(text: str, button_label: str = "Copy email"):
    """
    Render a browser-side copy button using navigator.clipboard.
    Falls back gracefully if clipboard access is blocked.
    """
    uid = uuid.uuid4().hex
    safe_text = json.dumps(text or "")
    components.html(
        f"""
        <div style="display:flex;align-items:center;gap:10px;margin:4px 0 10px 0;">
          <button
            id="copyBtn_{uid}"
            style="
              background:#0f62fe;color:white;border:none;border-radius:6px;
              padding:8px 12px;cursor:pointer;font-weight:600;
            "
          >
            {button_label}
          </button>
          <span id="copyStatus_{uid}" style="font-size:0.9rem;color:#444;"></span>
        </div>
        <script>
          const text_{uid} = {safe_text};
          const btn_{uid} = document.getElementById("copyBtn_{uid}");
          const status_{uid} = document.getElementById("copyStatus_{uid}");
          btn_{uid}.addEventListener("click", async () => {{
            try {{
              await navigator.clipboard.writeText(text_{uid});
              status_{uid}.textContent = "Copied to clipboard.";
              setTimeout(() => status_{uid}.textContent = "", 2500);
            }} catch (e) {{
              status_{uid}.textContent = "Copy failed (browser blocked). Please select and copy the text below.";
            }}
          }});
        </script>
        """,
        height=60,
    )

# ------------------------------------------------------------
# HELPERS
# ------------------------------------------------------------

def find_canonical_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None

def sanitize_sheet_name(name):
    name = re.sub(r'[\\/?*\[\]:]', '', name)
    return name[:31]

def build_mandatory_issues(df):
    issues = []
    missing_by_col = Counter()
    mandatory_cols = getattr(validator, "MANDATORY_COLS", [])
    for idx, row in df.iterrows():
        stock = row.get("stock_num", None)
        for col in mandatory_cols:
            if col in df.columns:
                if col == "color":
                    # If Color is missing, allow Fancy Color fields to satisfy the requirement.
                    color_missing = validator.is_empty_value(row.get("color", None))
                    fancy_present = (
                        not validator.is_empty_value(row.get("fancy_color_dominant_color", None))
                        or not validator.is_empty_value(row.get("fancy_color_intensity", None))
                    )
                    if not (color_missing and not fancy_present):
                        continue

                if validator.is_empty_value(row[col]) if col != "color" else True:
                    issues.append({
                        "Category": "Missing Mandatory",
                        "Stock No.": stock if not validator.is_empty_value(stock) else f"Row {idx + 2}",
                        "Issue Type": "Missing Value",
                        "Column": col,
                        "Value": row[col],
                        "Details": "Missing mandatory field",
                        "Row": idx + 2,
                    })
                    missing_by_col[col] += 1
    return issues, missing_by_col

def parse_invalid_value_strings(invalid_list, df):
    issues = []
    invalid_shape_values = set()
    invalid_color_values = set()
    invalid_by_col = Counter()

    pattern = re.compile(r"Row (\d+): Invalid '(.*)' in column '([^']+)'")

    for msg in invalid_list:
        m = pattern.match(msg)
        if not m:
            continue
        row_num = int(m.group(1))
        value = m.group(2)
        column = m.group(3)

        data_idx = row_num - 2
        if data_idx < 0 or data_idx >= len(df):
            continue
        row = df.iloc[data_idx]
        stock = row.get("stock_num", None)
        if validator.is_empty_value(stock):
            stock = f"Row {row_num}"

        issues.append({
            "Category": "Invalid Value",
            "Stock No.": stock,
            "Issue Type": "Invalid Value",
            "Column": column,
            "Value": value,
            "Details": "Value not in accepted list",
            "Row": row_num,
        })
        
        invalid_by_col[column] += 1

        if column == "shape":
            invalid_shape_values.add(value)
        elif column == "color":
            invalid_color_values.add(value)

    return issues, sorted(invalid_shape_values), sorted(invalid_color_values), invalid_by_col

def parse_numeric_invalid_strings(numeric_list, df):
    issues = []
    invalid_by_col = Counter()

    pattern = re.compile(r"Row (\d+): Invalid (?:carat value|price) '(.*)' in column '([^']+)'")

    for msg in numeric_list:
        m = pattern.match(msg)
        if not m:
            continue
        row_num = int(m.group(1))
        value = m.group(2)
        column = m.group(3)

        data_idx = row_num - 2
        if data_idx < 0 or data_idx >= len(df):
            continue
        row = df.iloc[data_idx]
        stock = row.get("stock_num", None)
        if validator.is_empty_value(stock):
            stock = f"Row {row_num}"

        if "carat" in column or "weight" in column or "size" in column:
            detail = "Carat/Weight must be greater than 0"
            category = "Invalid Value"
        else:
            detail = "Price must be greater than 0"
            category = "Price Issue"

        issues.append({
            "Category": category,
            "Stock No.": stock,
            "Issue Type": "Invalid Numeric Value",
            "Column": column,
            "Value": value,
            "Details": detail,
            "Row": row_num,
        })
        
        invalid_by_col[column] += 1

    return issues, invalid_by_col

def parse_url_issue_strings(url_list, df):
    issues = []
    pattern = re.compile(r"Row (\d+): ([^ ]+) → (.+?) → URL: (.+)")

    missing_image = 0
    missing_video = 0
    bad_video = 0
    bad_image = 0
    bad_cert = 0

    for msg in url_list:
        m = pattern.match(msg)
        if not m:
            continue
        row_num = int(m.group(1))
        col = m.group(2)
        status = m.group(3)
        url_value = m.group(4)

        data_idx = row_num - 2
        if data_idx < 0 or data_idx >= len(df):
            continue
        row = df.iloc[data_idx]
        stock = row.get("stock_num", None)
        if validator.is_empty_value(stock):
            stock = f"Row {row_num}"

        if "NOT PROVIDED" in status:
            issue_type = "Missing URL"
        else:
            issue_type = "URL Error"

        issues.append({
            "Category": "URL Issue",
            "Stock No.": stock,
            "Issue Type": issue_type,
            "Column": col,
            "URL": url_value,
            "Status": status,
            "Row": row_num,
        })

        if col == "image_url_1":
            if "NOT PROVIDED" in status:
                missing_image += 1
            else:
                bad_image += 1
        elif col == "video_url_1":
            if "NOT PROVIDED" in status:
                missing_video += 1
            else:
                bad_video += 1
        elif col == "cert_url_1":
            # Count any cert URL issue (missing, not working, unacceptable format)
            bad_cert += 1
            
    counts = {
        "missing_image": missing_image,
        "missing_video": missing_video,
        "bad_video": bad_video,
        "bad_image": bad_image,
        "bad_cert": bad_cert,
    }
    return issues, counts

def find_missing_cut_grade(df):
    issues = []
    count = 0
    cut_cols = [c for c in ["cut_grade", "cut"] if c in df.columns]
    if not cut_cols:
        return issues, 0

    col = cut_cols[0]

    def is_round_shape(shape_value):
        key = validator.normalize_value_key(shape_value)
        if not key:
            return False
        # Common variants: "Round", "Round Brilliant", "RB", "RD"
        return key.startswith("round") or key in {"rb", "rd"}

    for idx, row in df.iterrows():
        # Cut grade is mandatory only for Round stones.
        if "shape" not in df.columns or validator.is_empty_value(row.get("shape", None)):
            continue
        if not is_round_shape(row.get("shape", None)):
            continue

        stock = row.get("stock_num", None)
        if validator.is_empty_value(stock):
            stock = f"Row {idx + 2}"
        if validator.is_empty_value(row[col]):
            count += 1
            issues.append({
                "Category": "Missing Value",
                "Stock No.": stock,
                "Issue Type": "Missing Value",
                "Column": col,
                "Value": row[col],
                "Details": "Missing cut grade",
                "Row": idx + 2,
            })
    return issues, count

def build_price_mismatch_issues(df):
    issues = []
    count = 0

    def find_by_norm(candidates):
        for col in df.columns:
            n = validator.normalize_header_name(col)
            if n in candidates:
                return col
        return None

    # Look up by normalized header name so this works even if the supplier header
    # wasn't mapped to a canonical name by headers.xlsx.
    weight_col = find_by_norm({"carat", "weight", "carat_weight", "size"})
    ppc_col = find_by_norm({"price_per_carat", "ppc"})
    tsp_col = find_by_norm({"total_sales_price", "total_price", "total", "amount"})

    if not (weight_col and ppc_col and tsp_col):
        return issues, 0

    def to_float(x):
        if validator.is_empty_value(x):
            return None
        s = str(x).strip().replace(",", "")
        s = re.sub(r"[^\d.\-]", "", s)
        if not s or s in {"-", ".", "-."}:
            return None
        try:
            return float(s)
        except Exception:
            return None

    for idx, row in df.iterrows():
        stock = row.get("stock_num", None)
        if validator.is_empty_value(stock):
            stock = f"Row {idx + 2}"
        w = to_float(row.get(weight_col, None))
        ppc = to_float(row.get(ppc_col, None))
        tsp = to_float(row.get(tsp_col, None))

        if w is None or ppc is None or tsp is None:
            continue

        expected = round(w * ppc, 2)
        if abs(expected - tsp) > 0.01:
            count += 1
            issues.append({
                "Category": "Price Issue",
                "Stock No.": stock,
                "Issue Type": "Price Mismatch",
                "Column": validator.normalize_header_name(tsp_col) or str(tsp_col),
                "Value": row.get(tsp_col, None),
                "Details": f"Expected {expected} = {w} * {ppc}, got {tsp}",
                "Row": idx + 2,
            })

    return issues, count

def build_excel_report(structured_issues, df=None):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    section_map = {
        "stock_num": "1. Stock Number",
        "shape": "2. Shape",
        "weight": "3. Weight",
        "carat": "3. Weight",
        "carat_weight": "3. Weight",
        "size": "3. Weight",
        "color": "4. Color",
        "clarity": "5. Clarity",
        "image_url_1": "6. Image URL",
        "video_url_1": "7. Video URL",
        "cert_url_1": "8. Certificate URL",
        "price_per_carat": "9. Price",
        "total_sales_price": "9. Price",
        "cut": "10. Cut (Round)",
        "cut_grade": "10. Cut (Round)",
    }
    other_sheet = "11. Other Issues"
    
    issues_by_sheet = {}
    
    # Desired workbook tab order (avoid lexicographic "10." coming between "1." and "2.")
    sheet_order = [
        "Summary",
        "1. Stock Number",
        "2. Shape",
        "3. Weight",
        "4. Color",
        "5. Clarity",
        "6. Image URL",
        "7. Video URL",
        "8. Certificate URL",
        "9. Price",
        "10. Cut (Round)",
        other_sheet,
        "No Issues Found",
    ]

    all_sheet_names = set(section_map.values())
    all_sheet_names.add(other_sheet)
    for name in all_sheet_names:
        issues_by_sheet[name] = []
        
    for issue in structured_issues:
        column = issue.get("Column", "").lower()
        sheet_name = section_map.get(column, other_sheet)
        
        if issue["Issue Type"] == "Price Mismatch":
             sheet_name = "9. Price"

        issues_by_sheet[sheet_name].append(issue)

    # -----------------------------
    # Helper formatting functions
    # -----------------------------
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    # Darker border so it is clearly visible in Excel.
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    bad_cell_fill = PatternFill("solid", fgColor="FFC7CE")  # Excel-like light red

    issue_type_fills = {
        "Missing Value": PatternFill("solid", fgColor="F8D7DA"),           # light red
        "Missing URL": PatternFill("solid", fgColor="F8D7DA"),             # light red
        "Invalid Value": PatternFill("solid", fgColor="FFF3CD"),           # light yellow
        "Invalid Numeric Value": PatternFill("solid", fgColor="FFF3CD"),   # light yellow
        "URL Error": PatternFill("solid", fgColor="D1ECF1"),               # light blue
        "Price Mismatch": PatternFill("solid", fgColor="E2D9F3"),           # light purple
    }

    def auto_fit_columns(ws, max_width=65):
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells:
                val = cell.value
                if val is None:
                    continue
                s = str(val)
                if len(s) > max_len:
                    max_len = len(s)
            width = min(max(10, max_len + 2), max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def apply_borders(ws):
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = thin_border

    def style_table_sheet(ws):
        # Header styling
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 22
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Body styling + row color coding
        issue_type_col = None
        for i, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip().lower() == "issue type":
                issue_type_col = i
                break

        for r in range(2, ws.max_row + 1):
            issue_type = None
            if issue_type_col:
                issue_type = ws.cell(row=r, column=issue_type_col).value
            fill = issue_type_fills.get(str(issue_type), None)
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = body_alignment
                cell.border = thin_border
                if fill:
                    cell.fill = fill
                elif r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7F7F7")  # zebra striping

        auto_fit_columns(ws)
        apply_borders(ws)

    def col_letters_for(candidates):
        """
        Map canonical-ish columns to Excel column letters based on current df ordering.
        """
        if df is None or getattr(df, "empty", True):
            return ""
        letters = []
        for idx, c in enumerate(list(df.columns), start=1):
            n = validator.normalize_header_name(c)
            if n in candidates:
                letters.append(get_column_letter(idx))
        if not letters:
            return ""
        if len(letters) == 1:
            return f"Column {letters[0]}"
        return "Column " + " & ".join(letters)

    def summarize_column(col_key):
        """
        Return (missing_count, invalid_values_set, invalid_count) for a canonical column key.
        """
        missing_count = 0
        invalid_values = set()
        invalid_count = 0
        for issue in structured_issues:
            col = validator.normalize_header_name(issue.get("Column"))
            itype = issue.get("Issue Type") or ""
            if col != col_key:
                continue
            if itype == "Missing Value":
                missing_count += 1
            if itype in {"Invalid Value", "Invalid Numeric Value"}:
                invalid_count += 1
                v = issue.get("Value")
                if v is not None and str(v).strip() != "":
                    invalid_values.add(str(v).strip())
        return missing_count, invalid_values, invalid_count

    def summarize_urls(col_key):
        """
        Return missing_count and bad_count for URL columns.
        """
        missing_count = 0
        bad_count = 0
        for issue in structured_issues:
            col = validator.normalize_header_name(issue.get("Column"))
            if col != col_key:
                continue
            itype = issue.get("Issue Type") or ""
            if itype == "Missing URL":
                missing_count += 1
            elif itype == "URL Error":
                bad_count += 1
        return missing_count, bad_count

    def clip_list(values, limit=20):
        values = [v for v in values if v is not None and str(v).strip() != ""]
        values = list(dict.fromkeys(values))  # stable unique
        if len(values) <= limit:
            return ", ".join(values)
        return ", ".join(values[:limit]) + f" ... (+{len(values) - limit} more)"

    def build_summary_rows():
        """
        Build a client-facing summary table similar to the requested template.
        """
        rows = []
        sr = 1
        def compute_total_items():
            if df is None:
                return ""
            if getattr(df, "empty", True):
                return 0
            # Ignore fully blank rows (common in exported XLSX)
            total = 0
            for _, r in df.iterrows():
                if any(not validator.is_empty_value(v) for v in r.values):
                    total += 1
            return total

        total_items = compute_total_items()
        def compute_total_impacted_items():
            impacted = set()
            max_row = (len(df) + 1) if df is not None and not getattr(df, "empty", True) else None
            for issue in structured_issues:
                try:
                    r = int(issue.get("Row"))
                except Exception:
                    continue
                if r < 2:
                    continue
                if max_row is not None and r > max_row:
                    continue
                impacted.add(r)
            return len(impacted)

        impacted_items = compute_total_impacted_items()

        # Stock Number
        m_cnt, inv_vals, inv_cnt = summarize_column("stock_num")
        if m_cnt or inv_cnt:
            parts = ["Stock Number:"]
            if m_cnt:
                parts.append(f"- Stock number is missing for {m_cnt} item(s) in your inventory file.")
            if inv_vals:
                parts.append(f"- We are receiving the value(s) {clip_list(sorted(inv_vals))} under this column which is not accepted.")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"stock_num"}),
                "Issue": "\n".join(parts),
                "Critical": "Yes",
                "Impact": "Items will not be listed or searchable on the app",
                "Link": "Go to sheet",
                "Items Impacted": int(m_cnt + inv_cnt),
                "Resolution": "Please provide unique Stock Number for all items.",
                "Comments": "",
                "__sheet": "1. Stock Number",
            })
            sr += 1

        # Weight / Carat (critical)
        weight_keys = {"size", "carat", "weight", "carat_weight"}
        w_missing = 0
        w_invalid = 0
        w_invalid_vals = set()
        for issue in structured_issues:
            col = validator.normalize_header_name(issue.get("Column"))
            if col not in weight_keys:
                continue
            itype = issue.get("Issue Type") or ""
            if itype == "Missing Value":
                w_missing += 1
            if itype in {"Invalid Value", "Invalid Numeric Value"}:
                w_invalid += 1
                v = issue.get("Value")
                if v is not None and str(v).strip() != "":
                    w_invalid_vals.add(str(v).strip())

        if w_missing or w_invalid:
            parts = ["Weight:"]
            if w_missing:
                parts.append(f"- The value is missing for {w_missing} item(s) in your inventory file.")
            if w_invalid_vals:
                parts.append(f"- We are receiving the value(s) {clip_list(sorted(w_invalid_vals))} under this column which is not accepted.")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for(weight_keys),
                "Issue": "\n".join(parts),
                "Critical": "Yes",
                "Impact": "Items will not be listed or searchable on the app",
                "Link": "Go to sheet",
                "Items Impacted": int(w_missing + w_invalid),
                "Resolution": "Please provide correct Weight/Carat values (> 00).",
                "Comments": "",
                "__sheet": "3. Weight",
            })
            sr += 1

        # Shape
        m_cnt, inv_vals, inv_cnt = summarize_column("shape")
        if m_cnt or inv_cnt:
            parts = ["Shape:"]
            if m_cnt:
                parts.append(f"- The value is missing for {m_cnt} item(s) in your inventory file.")
            if inv_vals:
                parts.append(f"- We are receiving the value(s) {clip_list(sorted(inv_vals))} under this column which is not accepted.")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"shape"}),
                "Issue": "\n".join(parts),
                "Critical": "Yes",
                "Impact": "Items will not be listed or searchable on the app",
                "Link": "Go to sheet",
                "Items Impacted": int(m_cnt + inv_cnt),
                "Resolution": "Please provide accepted Shape values as per VDB standardized list.",
                "Comments": "",
                "__sheet": "2. Shape",
            })
            sr += 1

        # Clarity
        m_cnt, inv_vals, inv_cnt = summarize_column("clarity")
        if m_cnt or inv_cnt:
            parts = ["Clarity:"]
            if m_cnt:
                parts.append(f"- The value is missing for {m_cnt} item(s) in your inventory file.")
            if inv_vals:
                parts.append(f"- We are receiving the value(s) {clip_list(sorted(inv_vals))} under this column which is not accepted.")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"clarity"}),
                "Issue": "\n".join(parts),
                "Critical": "Yes",
                "Impact": "Items will not be listed or searchable on the app",
                "Link": "Go to sheet",
                "Items Impacted": int(m_cnt + inv_cnt),
                "Resolution": "Please provide accepted Clarity values for all affected items.",
                "Comments": "",
                "__sheet": "5. Clarity",
            })
            sr += 1

        # Color
        m_cnt, inv_vals, inv_cnt = summarize_column("color")
        if m_cnt or inv_cnt:
            parts = ["Color:"]
            if m_cnt:
                parts.append(f"- The value is missing for {m_cnt} item(s) in your inventory file.")
            if inv_vals:
                parts.append(f"- We are receiving the value(s) {clip_list(sorted(inv_vals))} under this column which is not accepted.")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"color", "fancy_color_dominant_color", "fancy_color_intensity"}),
                "Issue": "\n".join(parts),
                "Critical": "Yes",
                "Impact": "Items will not be listed or searchable on the app",
                "Link": "Go to sheet",
                "Items Impacted": int(m_cnt + inv_cnt),
                "Resolution": "Please provide White color under Color, and Fancy color/intensity under the respective Fancy Color columns.",
                "Comments": "",
                "__sheet": "4. Color",
            })
            sr += 1

        # Price + mismatches
        m_ppc, inv_ppc_vals, inv_ppc_cnt = summarize_column("price_per_carat")
        m_tsp, inv_tsp_vals, inv_tsp_cnt = summarize_column("total_sales_price")
        mismatch_cnt = sum(1 for i in structured_issues if i.get("Issue Type") == "Price Mismatch")
        if m_ppc or inv_ppc_cnt or m_tsp or inv_tsp_cnt or mismatch_cnt:
            parts = ["Price & Total Price:"]
            if m_ppc:
                parts.append(f"- Price per carat is missing for {m_ppc} item(s).")
            if m_tsp:
                parts.append(f"- Total sales price is missing for {m_tsp} item(s).")
            if inv_ppc_vals:
                parts.append(f"- Invalid Price per carat value(s): {clip_list(sorted(inv_ppc_vals))}.")
            if inv_tsp_vals:
                parts.append(f"- Invalid Total sales price value(s): {clip_list(sorted(inv_tsp_vals))}.")
            if mismatch_cnt:
                parts.append(f"- For {mismatch_cnt} item(s), Total Sales Price does not match (Carat x Price Per Carat).")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"price_per_carat", "total_sales_price", "total_price", "amount", "total"}),
                "Issue": "\n".join(parts),
                "Critical": "Yes",
                "Impact": "Items may be visible with incorrect or missing prices on the app",
                "Link": "Go to sheet",
                "Items Impacted": int(m_ppc + inv_ppc_cnt + m_tsp + inv_tsp_cnt + mismatch_cnt),
                "Resolution": "Please provide correct Price Per Carat and Total Sales Price for all affected items.",
                "Comments": "",
                "__sheet": "9. Price",
            })
            sr += 1

        # Image URLs
        m_img, bad_img = summarize_urls("image_url_1")
        if m_img or bad_img:
            parts = ["Image URLs:"]
            if m_img:
                parts.append(f"- Image URLs are missing for {m_img} item(s) in your feed.")
            if bad_img:
                parts.append(f"- {bad_img} image URL(s) are not working or invalid.")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"image_url_1"}),
                "Issue": "\n".join(parts),
                "Critical": "No",
                "Impact": "Items would be visible without images on the app",
                "Link": "Go to sheet",
                "Items Impacted": int(m_img + bad_img),
                "Resolution": "Please provide direct, public image URLs to show them on the app.",
                "Comments": "",
                "__sheet": "6. Image URL",
            })
            sr += 1

        # Video URLs
        m_vid, bad_vid = summarize_urls("video_url_1")
        if m_vid or bad_vid:
            parts = ["Video URLs:"]
            if m_vid:
                parts.append(f"- Video URLs are missing for {m_vid} item(s) in your feed.")
            if bad_vid:
                parts.append(f"- {bad_vid} video URL(s) are not working or invalid.")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"video_url_1"}),
                "Issue": "\n".join(parts),
                "Critical": "No",
                "Impact": "Videos would not be visible under the media section for these stocks",
                "Link": "Go to sheet",
                "Items Impacted": int(m_vid + bad_vid),
                "Resolution": "Please provide direct source links to show the videos on the app.",
                "Comments": "",
                "__sheet": "7. Video URL",
            })
            sr += 1

        # Certificate URLs
        m_cert, bad_cert = summarize_urls("cert_url_1")
        if m_cert or bad_cert:
            parts = ["Certificate Links:"]
            if m_cert:
                parts.append(f"- We are not getting the certificate URLs for {m_cert} item(s) in your feed.")
            if bad_cert:
                parts.append(f"- {bad_cert} certificate URL(s) are not working or unacceptable format (only .pdf/.jpg allowed).")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"cert_url_1"}),
                "Issue": "\n".join(parts),
                "Critical": "No",
                "Impact": "Certificate would not be visible on the app for these items",
                "Link": "Go to sheet",
                "Items Impacted": int(m_cert + bad_cert),
                "Resolution": "Provide direct certificate URLs (accepted formats: .pdf/.jpg/.jpeg).",
                "Comments": "",
                "__sheet": "8. Certificate URL",
            })
            sr += 1

        # Cut (Round only)
        cut_missing = sum(
            1 for i in structured_issues
            if (i.get("Issue Type") == "Missing Value" and (i.get("Column") or "").lower() in {"cut", "cut_grade"})
        )
        if cut_missing:
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"cut", "cut_grade"}),
                "Issue": f"Cut:\n- We are not getting Cut value for {cut_missing} item(s) which have Round shape in your feed.",
                "Critical": "No",
                "Impact": "Cut value would not be visible on the app for these items",
                "Link": "Go to sheet",
                "Items Impacted": int(cut_missing),
                "Resolution": "Please provide Cut value for Round stones to show it on the app.",
                "Comments": "",
                "__sheet": "10. Cut (Round)",
            })
            sr += 1

        # Availability
        a_m, a_inv_vals, a_inv_cnt = summarize_column("availability")
        if a_m or a_inv_cnt:
            parts = ["Availability:"]
            if a_m:
                parts.append(f"- We are not getting the Availability for {a_m} item(s) in your feed.")
            if a_inv_vals:
                parts.append(f"- We are receiving the value(s) {clip_list(sorted(a_inv_vals))} under this column which is not accepted.")
            rows.append({
                "Sr. No.": sr,
                "Issue Column": col_letters_for({"availability"}),
                "Issue": "\n".join(parts),
                "Critical": "No",
                "Impact": "Availability would not be visible/accurate on the app for these items",
                "Link": "Go to sheet",
                "Items Impacted": int(a_m + a_inv_cnt),
                "Resolution": "Please provide accepted Availability values for all affected items.",
                "Comments": "",
                "__sheet": other_sheet,
            })
            sr += 1

        # Totals (footer)
        rows.append({
            "Sr. No.": sr,
            "Issue Column": "",
            "Issue": "Total Impacted Items",
            "Critical": "",
            "Impact": "",
            "Link": "",
            "Items Impacted": impacted_items,
            "Resolution": "",
            "Comments": f"Total items in file: {total_items}",
            "__sheet": None,
        })

        return rows

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Summary sheet first (client-facing)
        _summary_rows = build_summary_rows()
        df_summary = pd.DataFrame(_summary_rows, columns=[
            "Sr. No.",
            "Issue Column",
            "Issue",
            "Critical",
            "Impact",
            "Link",
            "Items Impacted",
            "Resolution",
            "Comments",
            "__sheet",
        ])
        # Hide internal column
        df_summary_out = df_summary.drop(columns=["__sheet"], errors="ignore")
        df_summary_out.to_excel(writer, sheet_name="Summary", index=False)

        # Write detail sheets as impacted inventory rows (full rows), with an 'issues' column.
        # This makes it easy to see the full context of each problematic stone.
        if df is not None and not df.empty:
            def issue_to_data_idx(issue):
                try:
                    row_num = int(issue.get("Row"))
                except Exception:
                    return None
                return row_num - 2  # data starts at Excel row 2

            for sheet_name in sheet_order:
                if sheet_name in {"Summary", "No Issues Found"}:
                    continue
                if sheet_name not in issues_by_sheet:
                    continue
                if not issues_by_sheet[sheet_name]:
                    continue

                idxs = []
                issues_by_idx = {}
                for iss in issues_by_sheet[sheet_name]:
                    di = issue_to_data_idx(iss)
                    if di is None or di < 0 or di >= len(df):
                        continue
                    idxs.append(di)
                    issues_by_idx.setdefault(di, []).append(iss)

                idxs = sorted(set(idxs))
                if not idxs:
                    continue

                safe_sheet_name = sanitize_sheet_name(sheet_name)
                df_rows = df.iloc[idxs].copy()
                df_rows.insert(0, "source_row", [i + 2 for i in idxs])

                def issue_text(di):
                    parts = []
                    for iss in issues_by_idx.get(di, []):
                        col = iss.get("Column")
                        it = iss.get("Issue Type")
                        det = iss.get("Details") or iss.get("Status") or ""
                        parts.append(f"{it} [{col}]{(': ' + det) if det else ''}")
                    return "\n".join(parts)

                df_rows["issues"] = [issue_text(i) for i in idxs]
                df_rows.to_excel(writer, sheet_name=safe_sheet_name, index=False)

        if not structured_issues:
            df_empty = pd.DataFrame(columns=["Stock No.", "Issue Type", "Column", "Value", "Details", "Row"])
            df_empty.to_excel(writer, sheet_name="No Issues Found", index=False)

        # Apply formatting
        wb = writer.book
        # Summary styling
        ws_sum = wb["Summary"]
        ws_sum.freeze_panes = "A2"
        ws_sum.auto_filter.ref = ws_sum.dimensions
        ws_sum.row_dimensions[1].height = 22
        for cell in ws_sum[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        for r in range(2, ws_sum.max_row + 1):
            for c in range(1, ws_sum.max_column + 1):
                cell = ws_sum.cell(row=r, column=c)
                cell.alignment = body_alignment
                cell.border = thin_border
        auto_fit_columns(ws_sum, max_width=95)
        # Make Issue/Resolution columns wider
        if ws_sum.max_column >= 3:
            ws_sum.column_dimensions[get_column_letter(3)].width = 55
        if ws_sum.max_column >= 8:
            ws_sum.column_dimensions[get_column_letter(8)].width = 55

        # Add sheet jump links in "Link" column (F)
        link_col_idx = 6
        for i, row in enumerate(_summary_rows, start=2):  # + header row
            target = row.get("__sheet")
            if not target:
                continue
            if target not in wb.sheetnames:
                continue
            cell = ws_sum.cell(row=i, column=link_col_idx)
            cell.value = f"=HYPERLINK(\"#'{target}'!A1\",\"Go to sheet\")"
            cell.style = "Hyperlink"

        apply_borders(ws_sum)

        # Detail sheet styling (in deterministic order)
        for name in wb.sheetnames:
            if name in {"Summary", "No Issues Found"}:
                continue
            style_table_sheet(wb[name])

        # Highlight the specific bad cells in red (based on issues list)
        if df is not None and not df.empty:
            for sheet_name in sorted(issues_by_sheet.keys()):
                safe_sheet_name = sanitize_sheet_name(sheet_name)
                if safe_sheet_name not in wb.sheetnames:
                    continue
                if safe_sheet_name in {"Summary", "No Issues Found"}:
                    continue

                ws = wb[safe_sheet_name]
                # Header -> column index
                header_to_idx = {}
                for c in range(1, ws.max_column + 1):
                    hv = ws.cell(row=1, column=c).value
                    if hv is None:
                        continue
                    header_to_idx[str(hv).strip().lower()] = c

                src_col_idx = header_to_idx.get("source_row")
                if not src_col_idx:
                    continue

                # Build a map: source_row_value -> worksheet row
                row_lookup = {}
                for r in range(2, ws.max_row + 1):
                    row_lookup[ws.cell(row=r, column=src_col_idx).value] = r

                for iss in issues_by_sheet[sheet_name]:
                    try:
                        source_row = int(iss.get("Row"))
                    except Exception:
                        continue
                    ws_row = row_lookup.get(source_row)
                    if not ws_row:
                        continue
                    col_key = (iss.get("Column") or "").strip().lower()
                    if not col_key:
                        continue
                    cidx = header_to_idx.get(col_key)
                    if not cidx:
                        continue
                    ws.cell(row=ws_row, column=cidx).fill = bad_cell_fill

        # "No Issues Found" styling
        if "No Issues Found" in wb.sheetnames:
            ws = wb["No Issues Found"]
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 22
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            auto_fit_columns(ws)
            apply_borders(ws)
                
    buffer.seek(0)
    return buffer

def build_email_body(
    supplier_name,
    invalid_shape_values,
    invalid_color_values,
    missing_by_col,
    invalid_by_col,
    url_counts,
    cut_missing_count,
    price_mismatch_count,
    missing_stock_count,
):
    lines = []
    lines.append(f"Hi {supplier_name},")
    lines.append("")
    lines.append("Hope you're doing well.")
    lines.append("")
    lines.append("During a routine validation of your inventory on the VDB Marketplace, we identified a few issues that need your attention. Please find the details below:")
    lines.append("")

    section_num = 1

    if missing_stock_count > 0:
        lines.append(f"{section_num}. Stock Number")
        lines.append(f"- Stock number is missing for {missing_stock_count} item(s).")
        lines.append("")
        section_num += 1

    if invalid_shape_values or missing_by_col.get("shape") or invalid_by_col.get("shape"):
        lines.append(f"{section_num}. Shape")
        if missing_by_col.get("shape"):
            lines.append(f"- Shape is missing for {missing_by_col['shape']} item(s).")
        if invalid_shape_values:
            lines.append("- We found invalid shape values that do not match VDB's standardised shape list, for example:")
            for sh in invalid_shape_values:
                lines.append(f"  • {sh}")
        lines.append("")
        section_num += 1

    weight_col = None
    for cand in ["carat", "weight", "carat_weight", "size"]:
        if cand in missing_by_col or cand in invalid_by_col:
            weight_col = cand
            break

    if weight_col and (missing_by_col.get(weight_col) or invalid_by_col.get(weight_col)):
        lines.append(f"{section_num}. Weight")
        if missing_by_col.get(weight_col):
            lines.append(f"- Weight ({weight_col}) is missing for {missing_by_col[weight_col]} item(s).")
        if invalid_by_col.get(weight_col):
            lines.append(f"- Weight ({weight_col}) has invalid values (zero, negative, or not in accepted format) for {invalid_by_col[weight_col]} item(s).")
        lines.append("")
        section_num += 1

    if missing_by_col.get("color") or invalid_by_col.get("color") or invalid_color_values:
        lines.append(f"{section_num}. Color")
        if missing_by_col.get("color"):
            lines.append(f"- Color is missing for {missing_by_col['color']} item(s).")
        if invalid_color_values:
            lines.append("- We found invalid color values that do not match VDB's standardised color list, for example:")
            for clr in invalid_color_values:
                lines.append(f"  • {clr}")
        lines.append("")
        section_num += 1

    if missing_by_col.get("clarity") or invalid_by_col.get("clarity"):
        lines.append(f"{section_num}. Clarity")
        if missing_by_col.get("clarity"):
            lines.append(f"- Clarity is missing for {missing_by_col['clarity']} item(s).")
        if invalid_by_col.get("clarity"):
            lines.append(f"- Clarity has invalid values for {invalid_by_col['clarity']} item(s).")
        lines.append("")
        section_num += 1

    missing_image = missing_by_col.get("image_url_1", 0) + url_counts.get("missing_image", 0)
    if missing_image or url_counts.get("bad_image", 0):
        lines.append(f"{section_num}. Image URLs")
        if missing_image:
            lines.append(f"- Image URLs are missing for {missing_image} item(s).")
        if url_counts.get("bad_image", 0):
            lines.append(f"- {url_counts['bad_image']} image URL(s) are not working (HTTP errors).")
        lines.append("")
        section_num += 1

    missing_video = missing_by_col.get("video_url_1", 0) + url_counts.get("missing_video", 0)
    if missing_video or url_counts.get("bad_video", 0):
        lines.append(f"{section_num}. Video URLs")
        if missing_video:
            lines.append(f"- Video URLs are missing for {missing_video} item(s).")
        if url_counts.get("bad_video", 0):
            lines.append(f"- {url_counts['bad_video']} video URL(s) are not working (HTTP errors).")
        lines.append("")
        section_num += 1

    cert_issue_present = (
        missing_by_col.get("cert_url_1", 0)
        or url_counts.get("bad_cert", 0)
    )
    if cert_issue_present:
        lines.append(f"{section_num}. Certificate URLs")
        if missing_by_col.get("cert_url_1", 0) + url_counts.get("bad_cert", 0) > 0:
             lines.append(f"- Certificate URLs are missing for {missing_by_col['cert_url_1'] + url_counts['bad_cert']} item(s).")
        lines.append("")
        section_num += 1

    price_issue_present = (
        missing_by_col.get("price_per_carat")
        or missing_by_col.get("total_sales_price")
        or invalid_by_col.get("price_per_carat")
        or invalid_by_col.get("total_sales_price")
        or price_mismatch_count
    )

    if price_issue_present:
        lines.append(f"{section_num}. Price")
        if missing_by_col.get("price_per_carat"):
            lines.append(f"- Price per carat is missing for {missing_by_col['price_per_carat']} item(s).")
        if missing_by_col.get("total_sales_price"):
            lines.append(f"- Total sales price is missing for {missing_by_col['total_sales_price']} item(s).")
        if invalid_by_col.get("price_per_carat"):
            lines.append(f"- Price per carat has invalid values (zero, negative, or not in accepted format) for {invalid_by_col['price_per_carat']} item(s).")
        if invalid_by_col.get("total_sales_price"):
            lines.append(f"- Total sales price has invalid values (zero, negative, or not in accepted format) for {invalid_by_col['total_sales_price']} item(s).")
        if price_mismatch_count:
            lines.append(f"- For {price_mismatch_count} item(s), Total Sales Price does not match (Carat x Price Per Carat).")
        lines.append("")
        section_num += 1

    if cut_missing_count:
        lines.append(f"{section_num}. Other Issues (Cut Grade)")
        lines.append(f"- Cut grade information is missing for {cut_missing_count} item(s).")
        lines.append("")

    lines.append("A spreadsheet outlining the above items has been attached for your reference. We would appreciate it if you could make the necessary corrections at your earliest convenience.")
    lines.append("")
    lines.append("If you have any questions or need further clarification, feel free to reach out. We'll be happy to assist.")
    lines.append("")
    lines.append("Best Regards,")
    lines.append("VDB Marketplace Support Team")

    return "\n".join(lines)

# ------------------------------------------------------------
# INITIALIZE SESSION STATE
# ------------------------------------------------------------

if 'validation_complete' not in st.session_state:
    st.session_state.validation_complete = False
if 'validation_results' not in st.session_state:
    st.session_state.validation_results = None
if 'file_uploader_key' not in st.session_state:
    st.session_state.file_uploader_key = 0
if 'supplier_name_value' not in st.session_state:
    st.session_state.supplier_name_value = "Supplier"

# ------------------------------------------------------------
# FILE UPLOAD UI
# ------------------------------------------------------------

supplier_file = st.file_uploader(
    "Upload Supplier Inventory (.csv or .xlsx)", 
    type=["csv", "xlsx"],
    key=f"file_uploader_{st.session_state.file_uploader_key}"
)

col1, col2 = st.columns([3, 1])
with col1:
    supplier_name = st.text_input(
        "Supplier Name (for email)", 
        value=st.session_state.supplier_name_value,
        key="supplier_name_input"
    )
    if supplier_name != st.session_state.supplier_name_value:
        st.session_state.supplier_name_value = supplier_name

with col2:
    st.write("")
    st.write("")
    if st.button("🔄 Reset", help="Clear all results and start fresh", key="reset_button", type="secondary"):
        st.session_state.validation_complete = False
        st.session_state.validation_results = None
        st.session_state.last_file_name = None
        st.session_state.supplier_name_value = "Supplier"
        st.session_state.file_uploader_key += 1
        st.rerun()

if supplier_file and st.session_state.get('last_file_name') != supplier_file.name:
    st.session_state.validation_complete = False
    st.session_state.validation_results = None
    st.session_state.last_file_name = supplier_file.name

start_btn = st.button("Run Validation", type="primary")

# ------------------------------------------------------------
# MAIN FLOW
# ------------------------------------------------------------

if start_btn and supplier_file:

    rules_path = "headers.xlsx"
    if not os.path.exists(rules_path):
        st.error(f"Configuration error: The rules file ({rules_path}) was not found.")
        st.stop()
        
    st.info("📘 Loading rules…")
    try:
        header_map, canonical_set = validator.load_header_rules(rules_path)
        value_rules = validator.load_value_rules(rules_path)
        st.success("Rules loaded successfully.")
        
    except Exception as e:
        st.error(f"Failed to load rules. Error: {e}")
        st.stop()
    
    st.info("📄 Loading supplier inventory…")
    supplier_bytes = supplier_file.read()
    ext = supplier_file.name.split(".")[-1].lower()

    try:
        df, load_meta = validator.load_supplier_bytes(
            file_bytes=supplier_bytes,
            filename=supplier_file.name,
            header_map=header_map,
        )
    except Exception as e:
        st.error(
            "Failed to read the uploaded inventory file. "
            "If this is an Excel file, please ensure it is a valid, non-password-protected `.xlsx` "
            "(not `.xls` and not a renamed file). If you started from a `.csv`, you can upload the `.csv` directly."
        )
        st.exception(e)
        st.stop()
    if ext == "xlsx" and load_meta.get("sheet_name") is not None:
        st.caption(
            f"XLSX detected. Using sheet **{load_meta['sheet_name']}** with header row **{load_meta['header_row'] + 1}**."
        )

    st.success(f"Supplier file loaded: **{len(df)} rows**")

    progress = st.progress(0)
    status = st.empty()

    status.text("Normalizing headers…")
    df, unknown_headers = validator.normalize_headers(df, header_map)
    progress.progress(12)

    status.text("Checking mandatory fields…")
    missing_strings = validator.check_mandatory(df)
    mandatory_issues, missing_by_col = build_mandatory_issues(df)
    missing_stock_count = sum(1 for issue in mandatory_issues if issue["Column"] == "stock_num")
    progress.progress(25)

    status.text("Checking numeric ranges…")
    numeric_invalid_strings = validator.check_numeric_ranges(df)
    numeric_invalid_issues, numeric_invalid_by_col = parse_numeric_invalid_strings(numeric_invalid_strings, df)
    progress.progress(40)

    status.text("Validating values…")
    invalid_strings = validator.check_values(df, value_rules)
    invalid_issues, invalid_shape_values, invalid_color_values, invalid_by_col = parse_invalid_value_strings(invalid_strings, df)
    progress.progress(60)

    for col, count in numeric_invalid_by_col.items():
        invalid_by_col[col] += count

    status.text("Checking URLs… (fast mode)")
    url_strings = validator.check_all_urls(df)
    url_issues_struct, url_counts = parse_url_issue_strings(url_strings, df)
    progress.progress(75)

    status.text("Checking cut grade and price consistency…")
    cut_issues, cut_missing_count = find_missing_cut_grade(df)
    price_issues, price_mismatch_count = build_price_mismatch_issues(df)
    progress.progress(90)

    status.text("Building reports…")
    structured_issues = []
    structured_issues.extend(mandatory_issues)
    structured_issues.extend(numeric_invalid_issues)
    structured_issues.extend(invalid_issues)
    structured_issues.extend(url_issues_struct)
    structured_issues.extend(cut_issues)
    structured_issues.extend(price_issues)

    excel_buffer = build_excel_report(structured_issues, df=df)

    email_body = build_email_body(
        supplier_name=supplier_name,
        invalid_shape_values=invalid_shape_values,
        invalid_color_values=invalid_color_values,
        missing_by_col=missing_by_col,
        invalid_by_col=invalid_by_col,
        url_counts=url_counts,
        cut_missing_count=cut_missing_count,
        price_mismatch_count=price_mismatch_count,
        missing_stock_count=missing_stock_count,
    )
    
    progress.progress(100)
    status.text("✅ Validation completed!")

    st.session_state.validation_complete = True
    st.session_state.validation_results = {
        'df': df,
        'unknown_headers': unknown_headers,
        'missing_strings': missing_strings,
        'numeric_invalid_strings': numeric_invalid_strings,
        'invalid_strings': invalid_strings,
        'url_strings': url_strings,
        'structured_issues': structured_issues,
        'email_body': email_body,
        'invalid_shape_values': invalid_shape_values,
        'invalid_color_values': invalid_color_values,
        'missing_by_col': missing_by_col,
        'invalid_by_col': invalid_by_col,
        'url_counts': url_counts,
        'cut_missing_count': cut_missing_count,
        'price_mismatch_count': price_mismatch_count,
        'missing_stock_count': missing_stock_count,
        'excel_buffer': excel_buffer,
        'supplier_name': supplier_name,
    }
    
    st.rerun()

# ------------------------------------------------------------
# DISPLAY RESULTS (from session state if available)
# ------------------------------------------------------------

if st.session_state.validation_complete and st.session_state.validation_results:
    
    results = st.session_state.validation_results
    
    st.success("✅ Validation completed!")
    
    st.subheader("📌 Raw Validation Output")

    if results['unknown_headers']:
        st.warning("⚠ Unknown Headers Found")
        st.write(results['unknown_headers'])

    if results['missing_strings']:
        st.error("❌ Missing Mandatory Fields")
        st.write(results['missing_strings'])

    if results['numeric_invalid_strings']:
        st.error("❌ Invalid Numeric Values (Zero/Negative)")
        st.write(results['numeric_invalid_strings'])

    if results['invalid_strings']:
        st.error("❌ Invalid Values Found")
        st.write(results['invalid_strings'])

    filtered_url_strings = [s for s in results['url_strings'] if 'cert_url_1' not in s or 'NOT PROVIDED' in s]
    
    if filtered_url_strings:
        st.error("❌ URL Issues (Image/Video/Cert)")
        st.write(filtered_url_strings)
    else:
        st.success("✅ All URLs are working or missing.")

    st.subheader("📊 Download Detailed Spreadsheet")
    st.download_button(
        label="📥 Download validation_report.xlsx",
        data=results['excel_buffer'],
        file_name="validation_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_excel"
    )

    st.subheader("📧 Email Summary")

    render_copy_to_clipboard(results["email_body"], button_label="Copy email to clipboard")
    st.text_area("Email to Supplier", value=results['email_body'], height=400, key="email_text", label_visibility="collapsed")

